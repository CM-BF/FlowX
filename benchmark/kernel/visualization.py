"""
FileName: visualization.py
Description: visulization tools
Time: 2020/7/30 11:30
Project: GNN_benchmark
Author: Shurui Gui
"""

import matplotlib.pyplot as plt
import openpyxl
import os
import numpy as np
from definitions import ROOT_DIR
from cilog import create_logger
from os.path import join as opj
import os.path as osp
import datetime
import matplotlib.ticker as mtick


create_logger(sub_print=True)

SMALL_SIZE = 16
MEDIUM_SIZE = 18
BIGGER_SIZE = 20

plt.rc('font', size=BIGGER_SIZE)          # controls default text sizes
plt.rc('axes', titlesize=BIGGER_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=MEDIUM_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=SMALL_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title

GCN_excel = openpyxl.load_workbook(os.path.join(ROOT_DIR, 'quantitative_results', 'GCN_PL_prob.xlsx'))
GIN_excel = openpyxl.load_workbook(os.path.join(ROOT_DIR, 'quantitative_results', 'GIN_PL_prob.xlsx'))

GCN_sheets = GCN_excel.worksheets
GIN_sheets = GIN_excel.worksheets


# csfont = {'fontname':'Times New Roman'}
# allowed_methods = ['FlowMask', 'FlowShap-orig', 'FlowShap-plus']#['GradCAM', 'DeepLIFT', 'GNNExplainer', 'PGExplainer', 'GNN-GI', 'GNN-LRP', 'FlowShap-plus', 'PGMExplainer'] # 'FlowShap-orig', , 'FlowMask', 'SubgraphX'
allowed_methods = ['Gem']
name_map = {'FlowShap-orig': 'FlowX (shap)', 'FlowShap-plus': 'FlowX', 'FlowMask': 'FlowX (mask)'}
dataset_names = ['BA-Shapes', 'BA-LRP', 'ClinTox', 'Tox21', 'BBBP', 'BACE', 'Graph-SST2', 'ba_infe']
# allowed_datasets = ['BA-Shapes', 'BA-LRP', 'ClinTox', 'Tox21', 'BBBP', 'BACE', 'Graph-SST2']
allowed_datasets = ['BA-LRP', 'ClinTox', 'Tox21', 'BBBP', 'BACE']
table = [[['' for _ in range(len(dataset_names) + 1)] for _ in range(len(allowed_methods) + 1)] for _ in range(2)]
for i in range(1, len(allowed_datasets) + 1):
    table[0][0][i] = allowed_datasets[i - 1]
    table[1][0][i] = allowed_datasets[i - 1]

for i in range(1, len(allowed_methods) + 1):
    table[0][i][0] = allowed_methods[i - 1]
    table[1][i][0] = allowed_methods[i - 1]

marker = ['^--C0', 'v--C1', 's--C2', 'p--C3', 'o--C4', 'D--C5', '>--C6', 'P-C7', '8--C8', '<--C9', 'd--C8', '^--C0']
    #['GradCAM', 'DeepLIFT', 'GNNExplainer', 'GNN_GI', 'WalkEraser', 'GNN_LRP', 'FlowX']
# --- for dataset_loaders ---
start_time = f'{datetime.datetime.now().strftime("%Y%m%d_%H_%M_%S")}'
legend_flag = True
handles = []
labels = []
single_figure_setting = {'figsize': (4.4, 4), 'dpi': 300}
ALL_PLOTS = False
if ALL_PLOTS:
    fig, axs = plt.subplots(2, len(dataset_names), figsize=(28, 10), dpi=250)
for model_idx, model in enumerate(['GCN', 'GIN']):
    for dataset_idx in range(GCN_sheets[5].max_column - 1):
        dataset_name = dataset_names[dataset_idx]
        if dataset_name not in allowed_datasets:
            continue
        for metric_idx, metric in enumerate(['Fidelity']):
            if ALL_PLOTS:
                ax = axs[model_idx][dataset_idx]
            else:
                fig, ax = plt.subplots(**single_figure_setting)
            for method_idx in range(GCN_sheets[5].max_row - 1):
                method_name = GCN_sheets[5].cell(method_idx + 2, 1).value.replace('_', '-')
                if method_name not in allowed_methods:
                    continue
                if dataset_idx == 0 and method_name in ['SubgraphX']:
                    continue
                line_x = []
                line_y = []
                for sparsity_idx in range(4, len(GCN_sheets)):
                    sparsity = float(sparsity_idx + 1) / 10.0
                    if model == 'GCN':
                        y = GCN_sheets[sparsity_idx].cell(method_idx + 2, dataset_idx + 2).value
                        if y is None:
                            continue
                        y = y.split('/')
                    elif model == 'GIN':
                        y = GIN_sheets[sparsity_idx].cell(method_idx + 2, dataset_idx + 2).value
                        if y is None:
                            continue
                        y = y.split('/')
                    else:
                        raise Exception('key error')
                    if metric == 'Fidelity':
                        y = float(y[0])
                    elif metric == 'Fidelity-':
                        y = float(y[1])
                    elif metric == 'Fidelity+-':
                        y = float(y[0]) - float(y[1])
                    else:
                        raise Exception('key error')
                    line_x.append(sparsity)
                    line_y.append(y)
                line_x = np.array(line_x)
                line_y = np.array(line_y)
                label = name_map.get(method_name) or method_name
                handle = ax.plot(line_x, line_y, marker[method_idx], label=label, linewidth=2, markerfacecolor='white',
                        markersize=12)
                if dataset_idx == 2 and model == 'GCN':
                    labels.append(label)
                    handles.append(handle[0])


                table[model_idx][allowed_methods.index(method_name) + 1][allowed_datasets.index(dataset_name) + 1] = [np.mean(line_y), np.std(line_y)]

            # ax.set_ylim([-0.1, 1])
            ax.set_xlabel('Sparsity')
            ax.set_ylabel(metric)
            ax.set_title(f'{dataset_name} ({model})')
            ax.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.2f'))
            if ALL_PLOTS and legend_flag and dataset_idx == 1:
                ax.legend()
                legend_flag = False
            if not ALL_PLOTS:
                plt.subplots_adjust(left=0.21, bottom=0.155, right=0.99, top=0.92, wspace=.001, hspace=.001)
                saving_root = opj(ROOT_DIR, 'quantitative_results', 'plots', start_time)
                if not osp.exists(saving_root):
                    os.mkdir(saving_root)
                # plt.show()
                # exit()
                fig.savefig(opj(saving_root, f'{model}_{dataset_name}.pdf'), format='pdf')
                plt.close(fig)

# fig.suptitle(f'{model}')
# --- mark the largest and the second largest the bold and underline ---
MARK_LARGEST = False

for model_idx in range(len(table)):
    for dataset_idx in range(1, len(dataset_names) + 1):
        max = - float('inf')
        max_idx = None
        second_max = - float('inf')
        second_idx = None
        for method_idx in range(1, len(allowed_methods) + 1):
            cell_value = table[model_idx][method_idx][dataset_idx]
            if cell_value and cell_value[0] > max:
                max = cell_value[0]
                max_idx = method_idx

        for method_idx in range(1, len(allowed_methods) + 1):
            cell_value = table[model_idx][method_idx][dataset_idx]
            if cell_value and cell_value[0] > second_max and method_idx != max_idx:
                second_max = cell_value[0]
                second_idx = method_idx

        for method_idx in range(1, len(allowed_methods) + 1):
            cell_value = table[model_idx][method_idx][dataset_idx]
            if cell_value:
                if MARK_LARGEST:
                    if method_idx == max_idx:
                        table[model_idx][method_idx][dataset_idx] = f'\\textbf{{{cell_value[0]:.2f}}}\\textpm {cell_value[1]:.2f}'
                    elif method_idx == second_idx:
                        table[model_idx][method_idx][dataset_idx] = f'\\underline{{{cell_value[0]:.2f}}}\\textpm {cell_value[1]:.2f}'
                    else:
                        table[model_idx][method_idx][dataset_idx] = f'{cell_value[0]:.2f}\\textpm {cell_value[1]:.2f}'
                else:
                    table[model_idx][method_idx][dataset_idx] = f'{cell_value[0]:.2f}$\\pm$ {cell_value[1]:.2f}'


print(f'#T#{table[0]}')
print(f'#T#{table[1]}')

if ALL_PLOTS:
    fig.tight_layout()
    fig.savefig(opj(ROOT_DIR, 'quantitative_results', 'plots', f'{datetime.datetime.now().strftime("%Y%m%d_%H_%M_%S")}'))
else:
    legend_fig, ax = plt.subplots(**single_figure_setting)
    ax.axis('off')
    # print(handles)
    handles = handles[:4] + handles[7:] + handles[4:7]
    labels = labels[:4] + labels[7:] + labels[4:7]
    legend = ax.legend(handles, labels, loc=3, framealpha=1, frameon=False)
    # legend_fig.canvas.draw()
    # bbox = legend.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
    legend_fig.savefig(opj(saving_root, 'legend.pdf'), format='pdf')
    plt.close(legend_fig)
    # plt.show()
    # legend_fig.show()

# fig.show()